#!/usr/bin/env python
# coding: utf-8

# In[1]:


from bs4 import BeautifulSoup
import pandas as pd
import argparse
import os
import requests
import tarfile
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook, Workbook
import subprocess

def create_or_copy_workbook(template_path, new_workbook_path):
    try:
        # Try to load the existing workbook
        existing_workbook = load_workbook(new_workbook_path)
    except FileNotFoundError:
        # If the workbook doesn't exist, create a new one as a copy of the template
        existing_workbook = load_workbook(template_path)
        existing_workbook.save(new_workbook_path)
        return


# In[3]:


def fill_cell_suspension(workbook, biosample_metadata: pd.DataFrame):
    
    sheet = workbook['Cell suspension']
    values=sheet.values
    cellsuspdf=pd.DataFrame(values)
    cellsuspdf.columns=cellsuspdf.loc[0]
    cellsuspdf.drop(0, axis=0, inplace=True)
    cellsuspdf.reset_index(drop=True, inplace=True)
    
    dftomerge = cellsuspdf.loc[4:0].copy()
    dftomerge.reset_index(drop=True, inplace=True)

    dftomerge['CELL SUSPENSION ID (Required)']=biosample_metadata['Sample_ID']
    dftomerge['CELL SUSPENSION NAME']=biosample_metadata['Title']
    dftomerge['BIOSAMPLES ACCESSION']=biosample_metadata['BioSample_ID']
    dftomerge['GENUS SPECIES (Required)']=biosample_metadata['Organism']
    dftomerge['NCBI TAXON ID (Required)']='9606'
    merged_df = pd.concat([cellsuspdf.iloc[:4], dftomerge], ignore_index=True)
    
    return merged_df


# In[4]:


def fill_specimen(workbook, biosample_metadata: pd.DataFrame):
    #Specimen sheet

    sheet = workbook['Specimen from organism']
    values=sheet.values
    specimendf=pd.DataFrame(values)
    specimendf.columns=specimendf.loc[0]
    specimendf.drop(0, axis=0, inplace=True)
    specimendf.reset_index(drop=True, inplace=True)
    
    dftomerge = specimendf.loc[4:0].copy()
    dftomerge.reset_index(drop=True, inplace=True)

    dftomerge['SPECIMEN FROM ORGANISM ID (Required)']=biosample_metadata['Sample_ID']
   # dftomerge['ORGAN (Required)']=biosample_metadata['tissue']
    dftomerge['BIOSAMPLES ACCESSION']=biosample_metadata['BioSample_ID']
    dftomerge['GENUS SPECIES (Required)']=biosample_metadata['Organism']
    dftomerge['NCBI TAXON ID (Required)']='9606'
    merged_df = pd.concat([specimendf.iloc[:4], dftomerge], ignore_index=True)
    
    return merged_df


# In[5]:


def fill_library_prep(biosample_metadata: pd.DataFrame):
    workbook=load_workbook("library_protocol_template.xlsx")
    sheet=workbook['Library preparation protocol']
    values=sheet.values
    librarydf=pd.DataFrame(values)
    librarydf.columns=librarydf.loc[0]
    librarydf.drop(0, axis=0, inplace=True)
    librarydf.reset_index(drop=True, inplace=True)
    updatedf = librarydf.loc[0:3].copy()
    updatedf.reset_index(drop=True, inplace=True)


    for item in biosample_metadata['Library strategy'].unique().tolist():
        matching_items = librarydf['LIBRARY CONSTRUCTION METHOD (Required)'].str.contains(item, case=False, na=False)
        updatedf = pd.concat([updatedf, librarydf[matching_items]], ignore_index=True)
        updatedf.reset_index(drop=True, inplace=True)
    
    return updatedf


# In[6]:


def fill_sequencing_tab(workbook, biosample_metadata: pd.DataFrame):
   
    sheet=workbook['Sequencing protocol']
    values=sheet.values
    seqdf=pd.DataFrame(values)
    seqdf.columns=seqdf.loc[0]
    seqdf.drop(0, axis=0, inplace=True)
    seqdf.reset_index(drop=True, inplace=True)
    dftomerge = seqdf.loc[4:0].copy()
    dftomerge.reset_index(drop=True, inplace=True)
    
    dftomerge['INSTRUMENT MANUFACTURER AND MODEL (Required)']=biosample_metadata['Instrument model'].unique()
    dftomerge['SEQUENCING PROTOCOL NAME']=biosample_metadata['Instrument model'].unique()
    models=biosample_metadata['Instrument model'].unique()
    protocolids=[model.replace(" ", "_") for model in models]
    dftomerge['SEQUENCING PROTOCOL ID (Required)'] = protocolids
    
    merged_df = pd.concat([seqdf.iloc[:4], dftomerge], ignore_index=True)
    return merged_df


# In[8]:


def fill_supp_files(workbook, soup, cellsuspdf, biosample_metadata):

    sheet=workbook['Analysis file']
    values=sheet.values
    seqdf=pd.DataFrame(values)
    seqdf.columns=seqdf.loc[0]
    seqdf.drop(0, axis=0, inplace=True)
    seqdf.reset_index(drop=True, inplace=True)
    dftomerge = seqdf.loc[4:0].copy()
    dftomerge.reset_index(drop=True, inplace=True)

    
    data=soup.find_all('Supplementary-Data')
    files=[item.text.split("/")[-1].strip() for item in data]
    dftomerge['FILE NAME (Required)'] = files
    dftomerge['FILE FORMAT (Required)'] = [file.split(".")[-1] for file in files]
    dftomerge['FILE SOURCE']='GEO'
    
    file_prefixes = [file.split("_")[0] for file in files]
   
    condition = [file in cellsuspdf['CELL SUSPENSION ID (Required)'].tolist() for file in file_prefixes]

    matching_rows = dftomerge[condition]
    matching_rows = matching_rows.copy()
    matching_rows['CELL SUSPENSION ID (Required)'] = matching_rows['FILE NAME (Required)'].str.split('_').str[0]
    matching_rows.reset_index(inplace=True)

    dftomerge['CELL SUSPENSION ID (Required)']=matching_rows['CELL SUSPENSION ID (Required)']
    
    # Initialize a variable to store the matching file
    parentfilelist=[]
    # Iterate through the file_prefixes to find the first file that starts with "GSE"
    for file in files:
        if file.startswith("GSE"):
            parentfilelist.append(file)
        
    for item in parentfilelist:
        condition = dftomerge['FILE NAME (Required)'] == item
        dftomerge['CELL SUSPENSION ID (Required)'].loc[condition]='||'.join(biosample_metadata['Sample_ID'].tolist())
    
    merged_df = pd.concat([seqdf.iloc[:4], dftomerge], ignore_index=True)
 
    return merged_df
    


# In[9]:


def write_workbook(sheet_name: str, GSE: str, merged_df: pd.DataFrame):# Load the existing workbook

    workbook = load_workbook(f"{GSE}.xlsx")

    # Select the desired sheet to replace (e.g., "Sheet1")
    sheet = workbook[sheet_name]

    # Convert the DataFrame (merged_df) to rows
    data = dataframe_to_rows(merged_df.fillna(''), index=False, header=True)

    # Clear the existing contents of the sheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    # Write the new data to the sheet
    for idx, row_data in enumerate(data, 1):
        for col, value in enumerate(row_data, 1):
            sheet.cell(row=idx, column=col, value=value)

    # Save the workbook with the updated sheet
    workbook.save(f"{GSE}.xlsx")


# In[14]:


def main(GSE):

    subprocess.run(['python', 'extract-geo-metadata.py', GSE])
    # Usage example:
    template_path = 'hca_template.xlsx'
    new_workbook_path = f"{GSE}.xlsx"
    
    create_or_copy_workbook(template_path, new_workbook_path)
    
    
    biosample_metadata = pd.read_csv(f'{GSE}.tsv',sep='\t')
    
    # Load the existing workbook
    workbook = load_workbook("hca_template.xlsx")
    
    with open(f'{GSE}_family.xml', 'r', encoding='utf-8') as file:
        xml_data = file.read()
    soup = BeautifulSoup(xml_data, 'xml')

    specimen_df= fill_specimen(workbook, biosample_metadata)
    write_workbook('Specimen from organism', GSE, specimen_df)
    print(f"Wrote Specimen tab to {GSE}.xlsx")

    cellsuspdf = fill_cell_suspension(workbook, biosample_metadata)
    write_workbook('Cell suspension', GSE, cellsuspdf)
    print(f"Wrote Cell suspension tab to {GSE}.xlsx")
    
    libraryprepdf= fill_library_prep(biosample_metadata)
    write_workbook('Library preparation protocol', GSE, libraryprepdf)
    print(f"Wrote Library preparation protocol tab to {GSE}.xlsx")
    
    seqdf=fill_sequencing_tab(workbook, biosample_metadata)
    write_workbook('Sequencing protocol', GSE, seqdf)
    print(f"Wrote Sequencing protocol tab to {GSE}.xlsx")
    
    suppfiledf=fill_supp_files(workbook, soup, cellsuspdf, biosample_metadata)
    write_workbook('Analysis file', GSE, suppfiledf)
    print(f"Wrote Analysis file tab to {GSE}.xlsx")


# In[16]:

if __name__ == "__main__":
	parser = argparse.ArgumentParser(description="Generate HCA template sheet from GEO accession. For use when geo-to-hca script fails due to lack of SRA metadata.")
	parser.add_argument('GSE', help="GSE value")
	args = parser.parse_args()
	main(args.GSE)


# In[ ]:




